import openpyxl
import pyautogui
from cryptography.fernet import Fernet
import base64
import os

# Carregar a chave de criptografia
key = None
try:
    with open("encryption_key.txt", "rb") as key_file:
        key = key_file.read()
except FileNotFoundError:
    print("Chave de criptografia não encontrada. Por favor, execute o script de interface primeiro.")
    exit()

cipher_suite = Fernet(key)

# Função para descriptografar dados
def decrypt_data(encrypted_data):
    return cipher_suite.decrypt(encrypted_data).decode()

# Função para proteger dados sensíveis
def protect_sensitive_data(data):
    encrypted_data = cipher_suite.encrypt(data.encode())
    return encrypted_data

# Função para inserir dados no Excel
def inserir_dados_excel(protected_data):
    try:
        # Descriptografar os dados
        decrypted_data = decrypt_data(protected_data)
        cliente, produto, quantidade, categoria = decrypted_data.split(',')

        # Abrir o arquivo Excel
        workbook = openpyxl.load_workbook('vendas_de_produtos.xlsx')
        vendas_sheet = workbook['vendas']

        # Encontrar a última linha vazia
        last_empty_row = vendas_sheet.max_row + 1

        # Inserir os dados
        vendas_sheet.cell(row=last_empty_row, column=1).value = cliente
        vendas_sheet.cell(row=last_empty_row, column=2).value = produto
        vendas_sheet.cell(row=last_empty_row, column=3).value = quantidade
        vendas_sheet.cell(row=last_empty_row, column=4).value = categoria

        # Salvar o arquivo
        workbook.save('vendas_de_produtos.xlsx')

        print(f"Dados inseridos com sucesso para {cliente}")
    except Exception as e:
        print(f"Erro ao inserir dados: {str(e)}")

# Função para ler dados do Excel
def ler_dados_excel():
    try:
        workbook = openpyxl.load_workbook('vendas_de_produtos.xlsx')
        vendas_sheet = workbook['vendas']
        last_row = vendas_sheet.max_row
        
        protected_data = ""
        for row in range(2, last_row + 1):  # Começar na segunda linha (índice 2)
            cliente = vendas_sheet.cell(row=row, column=1).value
            produto = vendas_sheet.cell(row=row, column=2).value
            quantidade = vendas_sheet.cell(row=row, column=3).value
            categoria = vendas_sheet.cell(row=row, column=4).value
            
            # Criptografar os dados antes de retorná-los
            protected_data += f"{cliente},{produto},{quantidade},{categoria}\n"
        
        return protected_data.strip()
    except FileNotFoundError:
        print("Arquivo 'vendas_de_produtos.xlsx' não encontrado. Por favor, verifique o caminho e tente novamente.")
    except Exception as e:
        print(f"Erro ao ler dados do Excel: {str(e)}")

# Função para iniciar o programa
def main():
    # Ler os dados do Excel
    protected_data = ler_dados_excel()

    # Inserir os dados no sistema
    inserir_dados_excel(protected_data)

if __name__ == "__main__":
    main()

